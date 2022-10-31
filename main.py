from pyrogram import Client
import datetime, time
import os
import requests

from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

api_id = os.getenv('APIID')
api_hash = os.getenv('TOKEN')
app = Client('my_account', api_id, api_hash)

half_day = 84000

def plural_days():
    # Делаем правильное окончание слова
    day = ['день', 'дня', 'дней']
    n = days()
    if n % 10 == 1 and n % 100 != 11:
        p = 0
    elif 2 <= n % 10 <= 4 and (n % 100 < 10 or n % 100 >= 20):
        p = 1
    else:
        p = 2
    return day[p]

def days():
    # Высчитываем день
    time_has_passed = datetime.datetime.now() - list_date
    return time_has_passed.days

def photo_memes():
    #забираем мемес
    URL = 'https://meme-api.herokuapp.com/gimme'
    response = requests.get(URL)
    response = response.json()
    random_meme = response.get('url')
    return random_meme

def Parsing():
    # Отправляем сообщение
    wb = load_workbook('./data.xlsx')
    sheet = wb['Лист1']
    max = sheet.max_row
    max_kines = 2
    while max+1 != max_kines:   
        list = []
        for i in range(1, 4):
            list.append(sheet.cell(row=max_kines, column=i).value)
        global list_date
        list_date = list[2]
        Send_Mess(list)
        max_kines += 1

def Send_Mess(list):
    days_score = days()
    app.send_message(
        list[0],
        'Привет, я бот! Мой создатель всего-лишь хотел напомнить, ' \
        f'что ты ему должен {list[1]} ₽. Уже прошло {days_score} '\
        f'{plural_days()}. Ахуеть можно. Во всяком'\
        'случае держи мемес на английском :)'
    )
    app.send_photo(
        list[0],
        photo_memes()
    )

def main():
    app.start()
    while True:
        try:
            Parsing()
        finally:
            time.sleep(half_day)


if __name__ == '__main__':
    main()

print("te")
